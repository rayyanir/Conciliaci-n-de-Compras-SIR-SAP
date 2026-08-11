"""
reconciliation.py — Lógica de conciliación SAP vs SIR
Compatible con COMPRAS SAP.xlsx, COMPRAS SIR PEPSI.xlsx, COMPRAS SIR LARKIN.xls/.xlsx
"""
import struct, zlib, re
from xml.etree import ElementTree as ET
from html.parser import HTMLParser
from collections import defaultdict
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


# ── Parsers de archivos ────────────────────────────────────────────────────────

def _read_zip_entry(filepath, target):
    with open(filepath, 'rb') as f:
        data = f.read()
    pos = 0
    while True:
        idx = data.find(b'\x50\x4b\x03\x04', pos)
        if idx == -1:
            return None
        ver, flags, comp, mtime, mdate, crc, csz, usz, fnlen, exlen = \
            struct.unpack_from('<5H3I2H', data, idx + 4)
        name = data[idx + 30:idx + 30 + fnlen].decode('utf-8', 'replace')
        ds = idx + 30 + fnlen + exlen
        if name == target:
            return zlib.decompress(data[ds:ds + csz], -15) if comp == 8 else data[ds:ds + csz]
        pos = idx + 4


def _parse_xlsx_sheet(filepath, sheet_path='xl/worksheets/sheet1.xml'):
    ss_xml = _read_zip_entry(filepath, 'xl/sharedStrings.xml')
    ss = []
    if ss_xml:
        root = ET.fromstring(ss_xml)
        for si in root.findall(f'{{{NS}}}si'):
            ss.append(''.join(t.text or '' for t in si.iter(f'{{{NS}}}t')))
    sheet_xml = _read_zip_entry(filepath, sheet_path)
    if not sheet_xml:
        return []
    root = ET.fromstring(sheet_xml)
    rows = []
    for row_el in root.iter(f'{{{NS}}}row'):
        row = {}
        for c in row_el:
            col = re.match(r'([A-Z]+)', c.get('r', ''))
            if not col:
                continue
            col = col.group(1)
            t = c.get('t', '')
            v_el = c.find(f'{{{NS}}}v')
            if v_el is None or v_el.text is None:
                row[col] = None
            elif t == 's':
                row[col] = ss[int(v_el.text)] if int(v_el.text) < len(ss) else ''
            else:
                row[col] = v_el.text
        rows.append(row)
    return rows


class _TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self.cur_row = []
        self.cur_cell = None

    def handle_starttag(self, tag, attrs):
        if tag == 'tr':
            self.cur_row = []
        elif tag in ('td', 'th'):
            self.cur_cell = ''

    def handle_endtag(self, tag):
        if tag in ('td', 'th'):
            if self.cur_cell is not None:
                self.cur_row.append(self.cur_cell.strip())
            self.cur_cell = None
        elif tag == 'tr':
            if any(c.strip() for c in self.cur_row):
                self.rows.append(self.cur_row[:])

    def handle_data(self, data):
        if self.cur_cell is not None:
            self.cur_cell += data


def _parse_html_table(filepath):
    content = None
    for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'utf-16'):
        try:
            with open(filepath, encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, Exception):
            continue
    if content is None:
        with open(filepath, encoding='latin-1', errors='replace') as f:
            content = f.read()
    p = _TableParser()
    p.feed(content)
    return p.rows


# ── Helpers ────────────────────────────────────────────────────────────────────

def _last5(v):
    if v is None:
        return None
    d = re.sub(r'[^0-9]', '', str(v))
    return d[-5:] if len(d) >= 5 else (d if d else None)


def _parse_num(v, european=False):
    if v is None or str(v).strip() in ('-', '', '–'):
        return None
    s = str(v).strip()
    if european:
        s = s.replace('.', '').replace(',', '.')
    try:
        return round(float(s), 2)
    except Exception:
        return None


def _excel_date(v):
    try:
        return (date(1899, 12, 30) + timedelta(days=int(float(v)))).strftime('%d/%m/%Y')
    except Exception:
        return str(v) if v else ''


# ── Carga de archivos ──────────────────────────────────────────────────────────

def _get_sheet_paths(filepath):
    """Devuelve dict {nombre_hoja: path_xml} leyendo workbook.xml y sus relaciones."""
    rels_xml = _read_zip_entry(filepath, 'xl/_rels/workbook.xml.rels')
    rels = {}
    if rels_xml:
        root = ET.fromstring(rels_xml)
        for r in root:
            target = r.get('Target', '')
            if not target.startswith('worksheets/'):
                target = 'worksheets/' + target.split('worksheets/')[-1] if 'worksheets/' in target else target
            rels[r.get('Id')] = f'xl/{target}'
    wb_xml = _read_zip_entry(filepath, 'xl/workbook.xml')
    if not wb_xml:
        return {}
    root = ET.fromstring(wb_xml)
    sheets = {}
    for s in root.iter(f'{{{NS}}}sheet'):
        rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        sheets[s.get('name', '')] = rels.get(rid, 'xl/worksheets/sheet1.xml')
    return sheets


def _load_sap(path):
    """Lee el archivo SAP detectando automáticamente la hoja con datos de facturas.
    Evita parsear el archivo dos veces reutilizando los datos ya leídos."""
    sheet_paths = _get_sheet_paths(path)
    rows = []
    found = False
    # Recorrer hojas en orden; conservar la primera que tenga 'Referencia' en col D
    # o en su defecto la que tenga más de 100 filas (la de datos reales).
    fallback_rows = []
    for name, spath in sheet_paths.items():
        parsed = _parse_xlsx_sheet(path, spath)
        if parsed and 'D' in parsed[0] and 'Referencia' in str(parsed[0].get('D', '') or ''):
            rows = parsed
            found = True
            break
        if len(parsed) > len(fallback_rows):
            fallback_rows = parsed
    if not found:
        rows = fallback_rows if fallback_rows else _parse_xlsx_sheet(path, 'xl/worksheets/sheet1.xml')
    sap_pepsi, sap_larkin = [], []
    for r in rows[1:]:
        if not any(r.values()):
            continue
        ref = str(r.get('D', '') or '').strip()
        cc_ben = str(r.get('L', '') or '').strip()
        cc = cc_ben[:-2] if cc_ben.endswith('VE') else cc_ben
        monto = _parse_num(r.get('I'))
        proveedor = str(r.get('K', '') or '').strip()
        texto_cab = str(r.get('E', '') or '').strip()
        inv5 = _last5(ref)
        rec = {'cc': cc, 'ref': ref, 'inv5': inv5, 'monto': monto,
               'proveedor': proveedor, 'texto_cab': texto_cab}
        if 'LARKIN' in proveedor.upper():
            sap_larkin.append(rec)
        else:
            sap_pepsi.append(rec)
    return sap_pepsi, sap_larkin


def _load_sir_xlsx(path, sheet_path='xl/worksheets/sheet1.xml'):
    """Carga un archivo SIR en formato xlsx (PEPSI o LARKIN).
    Detecta columnas dinamicamente y excluye compras con devolucion asociada."""
    rows = _parse_xlsx_sheet(path, sheet_path)

    header_idx  = 0
    col_total   = 'L'
    col_dev     = 'K'
    col_fac     = 'D'
    col_fecha   = 'B'
    col_cod     = 'E'
    col_cod_dev = 'F'
    col_related = 'G'

    for i, r in enumerate(rows):
        if 'Centro de Costo' in str(r.get('A', '') or ''):
            header_idx = i
            for col, val in r.items():
                v = str(val or '').strip().lower()
                if v == 'total factura':
                    col_total = col
                elif v in ('devolucion', 'devolución'):
                    col_dev = col
                elif v == 'factura':
                    col_fac = col
                elif v == 'fecha':
                    col_fecha = col
                elif 'código' in v and 'compras' in v:
                    col_cod = col
                elif 'código' in v and 'devolucion' in v:
                    col_cod_dev = col
                elif 'relacionado' in v:
                    col_related = col
            break

    data_rows = rows[header_idx + 1:]

    # Paso 1: mapear CódigoCompras → monto total devuelto (valor absoluto).
    # Si hay varias devoluciones para la misma compra se acumulan.
    devolution_map = {}   # cod_compras -> monto devuelto acumulado
    for r in data_rows:
        cod_dev_val = str(r.get(col_cod_dev) or '').strip()
        related_val = str(r.get(col_related) or '').strip()
        if cod_dev_val and cod_dev_val != '-' and related_val and related_val != '-':
            dev_amt = _parse_num(r.get(col_total))   # valor negativo en el archivo
            if dev_amt is not None:
                devolution_map[related_val] = devolution_map.get(related_val, 0) + abs(dev_amt)

    # Paso 2: cargar registros aplicando la lógica de devoluciones:
    #   - Devolución total (dev >= monto original) → excluir la factura
    #   - Devolución parcial (dev < monto original) → incluir con monto ajustado
    # IMPORTANTE: actualizar last_cc ANTES de cualquier exclusión para que
    # las filas siguientes del mismo CC hereden el código correctamente,
    # incluso si la primera fila del grupo fue excluida por devolución total.
    result = []
    last_cc = None
    for r in data_rows:
        # Actualizar CC primero, antes de verificar exclusiones
        cc_raw = r.get('A')
        if cc_raw and not str(cc_raw).startswith('='):
            last_cc = str(cc_raw).strip()

        # Saltar filas de "Sub Total :" (por tienda) y "Total :" (gran total al
        # final de la hoja) — ambas usan la palabra "Total" en la col. relacionada.
        if 'total' in str(r.get(col_related, '') or '').strip().lower():
            continue
        # Saltar las propias filas de devolución
        cod_dev_val = str(r.get(col_cod_dev) or '').strip()
        if cod_dev_val and cod_dev_val != '-':
            continue

        cod   = str(r.get(col_cod) or '').strip()
        total = _parse_num(r.get(col_total))

        # Aplicar devolución si existe
        dev_amt = devolution_map.get(cod, 0)
        if dev_amt > 0 and total is not None:
            net = round(total - dev_amt, 2)
            if net <= 0.005:          # Devolución total → excluir
                continue
            total = net               # Devolución parcial → ajustar monto

        cc = last_cc
        if cc is None:
            continue
        fac   = str(r.get(col_fac) or '').strip()
        fac5  = _last5(fac)
        dev   = _parse_num(r.get(col_dev))
        fecha = _excel_date(r.get(col_fecha))
        result.append({'cc': cc, 'factura': fac, 'inv5': fac5, 'total': total,
                       'devolucion': dev, 'fecha': fecha, 'cod': cod})
    return result


def _load_sir_pepsi(path):
    sheet_paths = _get_sheet_paths(path)
    sheet_path = 'xl/worksheets/sheet1.xml'
    for name, spath in sheet_paths.items():
        sample = _parse_xlsx_sheet(path, spath)
        if any('Centro de Costo' in str(r.get('A', '') or '') for r in sample[:15]):
            sheet_path = spath
            break
    return _load_sir_xlsx(path, sheet_path)


def _load_sir_larkin(path):
    with open(path, 'rb') as f:
        magic = f.read(4)
    if magic[:2] == b'PK':
        sheet_paths = _get_sheet_paths(path)
        sheet_path = 'xl/worksheets/sheet1.xml'
        for name, spath in sheet_paths.items():
            sample = _parse_xlsx_sheet(path, spath)
            if any('Centro de Costo' in str(r.get('A', '') or '') for r in sample[:15]):
                sheet_path = spath
                break
        return _load_sir_xlsx(path, sheet_path)
    else:
        html_rows = _parse_html_table(path)
        result = []
        last_cc = None
        for r in html_rows[5:]:
            # Saltar "Sub Total :" y un posible gran "Total :" final de la hoja
            if 'total' in ' '.join(r).lower():
                continue
            if len(r) < 12:
                continue
            cc_raw = r[0].strip()
            if cc_raw and cc_raw != '-':
                last_cc = cc_raw
            cc = last_cc
            if cc is None:
                continue
            fac     = r[3].strip()
            fac5    = _last5(fac)
            total   = _parse_num(r[11], european=True)
            dev     = _parse_num(r[10], european=True)
            fecha   = r[1].strip()
            cod     = r[4].strip() if len(r) > 4 else ''
            cod_dev = r[5].strip() if len(r) > 5 else ''
            result.append({'cc': cc, 'factura': fac, 'inv5': fac5, 'total': total,
                           'devolucion': dev, 'fecha': fecha, 'cod': cod, 'cod_dev': cod_dev})
        return result


# ── Motor de comparación ───────────────────────────────────────────────────────

def _compare(sap_list, sir_list, tolerance, vendor_label):
    sir_lu = defaultdict(list)
    sir_no_inv = []
    for r in sir_list:
        if r['inv5']:
            sir_lu[(r['cc'], r['inv5'])].append(r)
        else:
            sir_no_inv.append(r)

    sap_lu = defaultdict(list)
    for r in sap_list:
        if r['inv5'] and r['cc']:
            sap_lu[(r['cc'], r['inv5'])].append(r)

    sir_keys = set(sir_lu)
    sap_keys = set(sap_lu)
    matches_ok, matches_diff, only_sir, only_sap, sir_no_fac = [], [], [], [], []

    for key in sir_keys & sap_keys:
        for sr in sir_lu[key]:
            for sp in sap_lu[key]:
                sa, spa = sr['total'], sp['monto']
                diff = abs(sa - spa) if (sa is not None and spa is not None) else None
                row = {'vendor': vendor_label, 'cc': key[0], 'inv5': key[1],
                       'sir_factura': sr['factura'], 'sap_ref': sp['ref'],
                       'sir_total': sa, 'sap_monto': spa,
                       'diferencia': round(spa - sa, 2) if diff is not None else None,
                       'sir_fecha': sr['fecha'], 'sir_cod': sr.get('cod', ''),
                       'proveedor': sp['proveedor'], 'tolerancia': tolerance}
                (matches_ok if (diff is not None and diff <= tolerance) else matches_diff).append(row)

    for key in sir_keys - sap_keys:
        for sr in sir_lu[key]:
            only_sir.append({'vendor': vendor_label, 'cc': key[0], 'inv5': key[1],
                             'sir_factura': sr['factura'], 'sir_total': sr['total'],
                             'sir_fecha': sr['fecha'], 'sir_cod': sr.get('cod', '')})

    matched_no_inv = set()
    for key in sap_keys - sir_keys:
        for sp in sap_lu[key]:
            if sp['monto'] is not None and sp['monto'] < 0:
                continue
            paired = False
            for idx, sr in enumerate(sir_no_inv):
                if idx in matched_no_inv:
                    continue
                if sr['cc'] != sp['cc']:
                    continue
                sa, spa = sr['total'], sp['monto']
                diff = abs(sa - spa) if (sa is not None and spa is not None) else None
                if diff is not None and diff <= tolerance:
                    matched_no_inv.add(idx)
                    matches_ok.append({'vendor': vendor_label, 'cc': sp['cc'], 'inv5': sp['inv5'],
                                       'sir_factura': '(sin N° SIR)', 'sap_ref': sp['ref'],
                                       'sir_total': sa, 'sap_monto': spa,
                                       'diferencia': round(spa - sa, 2),
                                       'sir_fecha': sr.get('fecha', ''), 'sir_cod': sr.get('cod', ''),
                                       'proveedor': sp['proveedor'], 'tolerancia': tolerance})
                    paired = True
                    break
            if not paired:
                only_sap.append({'vendor': vendor_label, 'cc': sp['cc'], 'inv5': sp['inv5'],
                                 'ref': sp['ref'], 'sap_monto': sp['monto'],
                                 'proveedor': sp['proveedor'], 'texto_cab': sp['texto_cab']})

    for idx, r in enumerate(sir_no_inv):
        if idx not in matched_no_inv:
            nf = r.copy()
            nf['vendor'] = vendor_label
            sir_no_fac.append(nf)

    return matches_ok, matches_diff, only_sir, only_sap, sir_no_fac


# ── API pública ────────────────────────────────────────────────────────────────

def run_reconciliation(sap_path, pepsi_path, larkin_path):
    sap_pepsi, sap_larkin = _load_sap(sap_path)
    sir_pepsi = _load_sir_pepsi(pepsi_path)
    sir_larkin = _load_sir_larkin(larkin_path)
    p_ok, p_diff, p_sir, p_sap, p_nofac = _compare(sap_pepsi, sir_pepsi, 1.0, 'PEPSI')
    l_ok, l_diff, l_sir, l_sap, l_nofac = _compare(sap_larkin, sir_larkin, 5.0, 'LARKIN')
    return {
        'summary': {
            'pepsi':  {'ok': len(p_ok), 'diff': len(p_diff), 'only_sir': len(p_sir), 'only_sap': len(p_sap), 'no_fac': len(p_nofac)},
            'larkin': {'ok': len(l_ok), 'diff': len(l_diff), 'only_sir': len(l_sir), 'only_sap': len(l_sap), 'no_fac': len(l_nofac)}
        },
        'matches_ok':   p_ok   + l_ok,
        'matches_diff': p_diff + l_diff,
        'only_sir':     p_sir  + l_sir,
        'only_sap':     p_sap  + l_sap,
        'sir_no_fac':   p_nofac + l_nofac
    }


def _sum_sap_by_store(sap_pepsi, sap_larkin):
    """Suma monto SAP (todas las compras, incl. notas de credito en negativo) por tienda/CC."""
    totals = defaultdict(float)
    for rec in sap_pepsi + sap_larkin:
        cc = rec.get('cc')
        monto = rec.get('monto')
        if cc and monto is not None:
            totals[cc] += monto
    return dict(totals)


def _sum_sir_xlsx_by_store(path, sheet_path='xl/worksheets/sheet1.xml'):
    """Suma cruda de la columna 'Total factura' por tienda/CC, sin logica de
    emparejamiento de facturas. Las notas de credito/devoluciones ya vienen
    con signo negativo en el archivo, por lo que se restan automaticamente."""
    rows = _parse_xlsx_sheet(path, sheet_path)

    header_idx  = 0
    col_total   = 'L'
    col_related = 'G'

    for i, r in enumerate(rows):
        if 'Centro de Costo' in str(r.get('A', '') or ''):
            header_idx = i
            for col, val in r.items():
                v = str(val or '').strip().lower()
                if v == 'total factura':
                    col_total = col
                elif 'relacionado' in v:
                    col_related = col
            break

    data_rows = rows[header_idx + 1:]
    totals = defaultdict(float)
    last_cc = None
    for r in data_rows:
        cc_raw = r.get('A')
        if cc_raw and not str(cc_raw).startswith('='):
            last_cc = str(cc_raw).strip()
        # Saltar "Sub Total :" y el gran "Total :" final de la hoja
        if 'total' in str(r.get(col_related, '') or '').strip().lower():
            continue
        cc = last_cc
        if cc is None:
            continue
        total = _parse_num(r.get(col_total))
        if total is not None:
            totals[cc] += total
    return dict(totals)


def _sum_sir_html_by_store(path):
    """Igual que _sum_sir_xlsx_by_store pero para el formato HTML (.xls) de LARKIN."""
    html_rows = _parse_html_table(path)
    totals = defaultdict(float)
    last_cc = None
    for r in html_rows[5:]:
        # Saltar "Sub Total :" y un posible gran "Total :" final de la hoja
        if 'total' in ' '.join(r).lower():
            continue
        if len(r) < 12:
            continue
        cc_raw = r[0].strip()
        if cc_raw and cc_raw != '-':
            last_cc = cc_raw
        cc = last_cc
        if cc is None:
            continue
        total = _parse_num(r[11], european=True)
        if total is not None:
            totals[cc] += total
    return dict(totals)


def _find_cc_sheet_path(path):
    """Ubica la hoja que contiene 'Centro de Costo' en la col A (formato SIR xlsx)."""
    sheet_paths = _get_sheet_paths(path)
    sheet_path = 'xl/worksheets/sheet1.xml'
    for name, spath in sheet_paths.items():
        sample = _parse_xlsx_sheet(path, spath)
        if any('Centro de Costo' in str(r.get('A', '') or '') for r in sample[:15]):
            sheet_path = spath
            break
    return sheet_path


def get_store_totals(sap_path, pepsi_path, larkin_path, tolerance=0.01):
    """Totalizacion simple por tienda: suma TODAS las compras registradas en cada
    sistema (SAP, SIR Pepsi, SIR Larkin) restando notas de credito/devoluciones,
    SIN validar si la factura existe en la otra plataforma. Se usa para detectar
    tiendas donde el total global no cuadra entre sistemas."""
    sap_pepsi, sap_larkin = _load_sap(sap_path)
    total_sap = _sum_sap_by_store(sap_pepsi, sap_larkin)
    total_sap_pepsi  = _sum_sap_by_store(sap_pepsi, [])
    total_sap_larkin = _sum_sap_by_store([], sap_larkin)

    total_pepsi = _sum_sir_xlsx_by_store(pepsi_path, _find_cc_sheet_path(pepsi_path))

    with open(larkin_path, 'rb') as f:
        magic = f.read(4)
    if magic[:2] == b'PK':
        total_larkin = _sum_sir_xlsx_by_store(larkin_path, _find_cc_sheet_path(larkin_path))
    else:
        total_larkin = _sum_sir_html_by_store(larkin_path)

    all_ccs = sorted(set(total_sap) | set(total_pepsi) | set(total_larkin))
    rows = []
    for cc in all_ccs:
        sap_v        = round(total_sap.get(cc, 0.0), 2)
        sap_pepsi_v  = round(total_sap_pepsi.get(cc, 0.0), 2)
        sap_larkin_v = round(total_sap_larkin.get(cc, 0.0), 2)
        pepsi_v      = round(total_pepsi.get(cc, 0.0), 2)
        larkin_v     = round(total_larkin.get(cc, 0.0), 2)
        diff         = round(sap_v - (pepsi_v + larkin_v), 2)
        diff_pepsi   = round(sap_pepsi_v - pepsi_v, 2)
        diff_larkin  = round(sap_larkin_v - larkin_v, 2)
        rows.append({
            'cc': cc,
            'total_sap': sap_v,
            'total_sap_pepsi': sap_pepsi_v,
            'total_sap_larkin': sap_larkin_v,
            'total_pepsi': pepsi_v,
            'total_larkin': larkin_v,
            'diferencia': diff,
            'con_diferencia': abs(diff) > tolerance,
            'diferencia_pepsi': diff_pepsi,
            'con_diferencia_pepsi': abs(diff_pepsi) > tolerance,
            'diferencia_larkin': diff_larkin,
            'con_diferencia_larkin': abs(diff_larkin) > tolerance,
        })
    return rows


def get_all_cost_centers(results):
    ccs = set()
    for lst in [results['matches_ok'], results['matches_diff'],
                results['only_sir'], results['only_sap'], results['sir_no_fac']]:
        for r in lst:
            if r.get('cc'):
                ccs.add(r['cc'])
    return sorted(ccs)


# ── Generación del reporte Excel ───────────────────────────────────────────────

def generate_excel(results, output_path, period='Mayo 2026', store_totals=None):
    thin = Side(style='thin', color='BFBFBF')
    B = Border(left=thin, right=thin, top=thin, bottom=thin)

    def mk_fill(hex_): return PatternFill('solid', start_color=hex_)
    def mk_font(**kw): return Font(name='Arial', **kw)

    H_FILL = mk_fill('1F3864'); H_FONT = mk_font(bold=True, color='FFFFFF', size=11)
    H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    OK_F = mk_fill('E2EFDA'); WARN_F = mk_fill('FFF2CC')
    ERR_F = mk_fill('FCE4D6'); BLUE_F = mk_fill('DAEEF3')
    DIFF_F = mk_fill('FFD966'); TOT_F = mk_fill('2F5496')
    P_HDR = mk_fill('1F5C99'); L_HDR = mk_fill('7B2D8B')
    P_ROW = mk_fill('E9F0FB'); L_ROW = mk_fill('F5E6FA')
    P_ROW2 = mk_fill('D0E4F7'); L_ROW2 = mk_fill('EED5FA')
    GRAY_F = mk_fill('F2F2F2')

    def hdr(ws, row, cols, fill=None):
        for c in range(1, cols + 1):
            cell = ws.cell(row, c)
            cell.font = H_FONT; cell.fill = fill or H_FILL
            cell.alignment = H_ALIGN; cell.border = B

    def brow(ws, row_i, cols, fill):
        for c in range(1, cols + 1):
            cell = ws.cell(row_i, c)
            cell.font = mk_font(size=10); cell.border = B
            if fill: cell.fill = fill
            cell.alignment = Alignment(
                horizontal='right' if isinstance(cell.value, (int, float)) else 'left')

    def tot_row(ws, row_i, cols, sum_cols):
        ws.cell(row_i, 1, 'TOTAL')
        for c in sum_cols:
            ws.cell(row_i, c, f'=SUM({get_column_letter(c)}3:{get_column_letter(c)}{row_i - 1})')
        for c in range(1, cols + 1):
            ws.cell(row_i, c).font = mk_font(bold=True, color='FFFFFF', size=10)
            ws.cell(row_i, c).fill = TOT_F; ws.cell(row_i, c).border = B

    def title(ws, text, cols, hex_):
        ws.row_dimensions[1].height = 36
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        ws['A1'] = text
        ws['A1'].font = mk_font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = mk_fill(hex_)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    def col_w(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def section_hdr(ws, row_i, text, cols, fill):
        ws.merge_cells(f'A{row_i}:{get_column_letter(cols)}{row_i}')
        ws.cell(row_i, 1, text)
        ws.cell(row_i, 1).font = mk_font(bold=True, color='FFFFFF', size=11)
        ws.cell(row_i, 1).fill = fill
        ws.cell(row_i, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_i].height = 20

    p_ok    = [r for r in results['matches_ok']   if r['vendor'] == 'PEPSI']
    l_ok    = [r for r in results['matches_ok']   if r['vendor'] == 'LARKIN']
    p_diff  = [r for r in results['matches_diff'] if r['vendor'] == 'PEPSI']
    l_diff  = [r for r in results['matches_diff'] if r['vendor'] == 'LARKIN']
    p_sir   = [r for r in results['only_sir']     if r['vendor'] == 'PEPSI']
    l_sir   = [r for r in results['only_sir']     if r['vendor'] == 'LARKIN']
    p_sap   = [r for r in results['only_sap']     if r['vendor'] == 'PEPSI']
    l_sap   = [r for r in results['only_sap']     if r['vendor'] == 'LARKIN']
    p_nofac = [r for r in results['sir_no_fac']   if r['vendor'] == 'PEPSI']
    l_nofac = [r for r in results['sir_no_fac']   if r['vendor'] == 'LARKIN']

    wb = Workbook()

    # RESUMEN
    ws = wb.active; ws.title = 'RESUMEN'
    ws.row_dimensions[1].height = 44
    ws.merge_cells('A1:E1')
    ws['A1'] = f'CONCILIACIÓN SAP vs SIR — PEPSI & LARKIN — {period.upper()}'
    ws['A1'].font = mk_font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = mk_fill('1F3864')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.append(['Proveedor', 'Resultado', 'Descripción', 'Registros', 'Tolerancia ($)'])
    hdr(ws, 2, 5); ws.row_dimensions[2].height = 22

    def s_row(ws, vendor, lbl, desc, cnt, tol, fill, vfill):
        r = ws.max_row + 1
        ws.cell(r, 1, vendor); ws.cell(r, 2, lbl); ws.cell(r, 3, desc)
        ws.cell(r, 4, cnt); ws.cell(r, 5, tol)
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.font = mk_font(size=11, bold=(c == 4))
            cell.fill = vfill if c == 1 else fill; cell.border = B
            cell.alignment = Alignment(horizontal='center' if c in (1, 4, 5) else 'left', vertical='center')
        ws.row_dimensions[r].height = 20

    for lbl, desc, cnt, fill in [
        ('OK', 'Diferencia <= $1', len(p_ok), OK_F),
        ('Diferencia', 'Diferencia > $1', len(p_diff), WARN_F),
        ('Solo en SIR', 'En SIR, no en SAP', len(p_sir), ERR_F),
        ('Solo en SAP', 'En SAP, no en SIR (monto >= $0)', len(p_sap), BLUE_F)]:
        s_row(ws, 'PEPSI', lbl, desc, cnt, '$1.00', fill, P_HDR)
    for lbl, desc, cnt, fill in [
        ('OK', 'Diferencia <= $5', len(l_ok), OK_F),
        ('Diferencia', 'Diferencia > $5', len(l_diff), WARN_F),
        ('Solo en SIR', 'En SIR, no en SAP', len(l_sir), ERR_F),
        ('Solo en SAP', 'En SAP, no en SIR (monto >= $0)', len(l_sap), BLUE_F)]:
        s_row(ws, 'LARKIN', lbl, desc, cnt, '$5.00', fill, L_HDR)
    col_w(ws, [12, 20, 46, 12, 14])

    # DIFERENCIAS
    ws2 = wb.create_sheet('Diferencias de Monto'); NC = 8
    title(ws2, f'FACTURAS CON DIFERENCIA DE MONTO POR ENCIMA DE LA TOLERANCIA — {period.upper()}', NC, 'BF8F00')
    ws2.append(['Proveedor','Centro de Costo','Ult. 5 Fac.','Factura SIR','Fecha SIR','Monto SIR ($)','Monto SAP ($)','Diferencia ($)'])
    hdr(ws2, 2, NC)
    for vendor, rows, ra, rb, hfill in [('PEPSI', p_diff, P_ROW, P_ROW2, P_HDR), ('LARKIN', l_diff, L_ROW, L_ROW2, L_HDR)]:
        if rows:
            section_hdr(ws2, ws2.max_row+1, f'-- {vendor}  (tolerancia {"$1" if vendor=="PEPSI" else "$5"}) --', NC, hfill)
            for i, r in enumerate(sorted(rows, key=lambda x: (x['cc'], x['inv5'])), 1):
                ws2.append([r['vendor'],r['cc'],r['inv5'],r['sir_factura'],r['sir_fecha'],r['sir_total'],r['sap_monto'],r['diferencia']])
                brow(ws2, ws2.max_row, NC, ra if i%2==0 else rb)
                ws2.cell(ws2.max_row, 8).fill = DIFF_F
                ws2.cell(ws2.max_row, 8).font = mk_font(size=10, bold=True)
    tot_row(ws2, ws2.max_row+1, NC, [6, 7, 8])
    col_w(ws2, [10, 16, 12, 16, 12, 14, 14, 14])

    # SOLO EN SIR
    ws3 = wb.create_sheet('Solo en SIR'); NC3 = 6
    title(ws3, f'FACTURAS EN SIR QUE NO ESTAN EN SAP — {period.upper()}', NC3, 'C00000')
    ws3.append(['Proveedor','Centro de Costo','Ult. 5 Fac.','Factura SIR','Fecha SIR','Monto SIR ($)'])
    hdr(ws3, 2, NC3)
    for vendor, rows, ra, rb, hfill in [('PEPSI', p_sir, ERR_F, mk_fill('FCF0ED'), P_HDR), ('LARKIN', l_sir, L_ROW, L_ROW2, L_HDR)]:
        if rows:
            section_hdr(ws3, ws3.max_row+1, f'-- {vendor} --', NC3, hfill)
            for i, r in enumerate(sorted(rows, key=lambda x: (x['cc'], str(x.get('sir_factura','')))), 1):
                ws3.append([r['vendor'],r['cc'],r['inv5'],r.get('sir_factura',''),r.get('sir_fecha',''),r.get('sir_total')])
                brow(ws3, ws3.max_row, NC3, ra if i%2==0 else rb)
    tot_row(ws3, ws3.max_row+1, NC3, [6])
    col_w(ws3, [10, 16, 12, 18, 12, 14])

    # SOLO EN SAP
    ws4 = wb.create_sheet('Solo en SAP'); NC4 = 6
    title(ws4, f'FACTURAS EN SAP QUE NO ESTAN EN SIR — {period.upper()}', NC4, '1F497D')
    ws4.append(['Proveedor','Centro de Costo','Ref SAP','Ult. 5 Fac.','Monto SAP ($)','Descripcion'])
    hdr(ws4, 2, NC4)
    for vendor, rows, ra, rb, hfill in [('PEPSI', p_sap, BLUE_F, mk_fill('E4EFF7'), P_HDR), ('LARKIN', l_sap, L_ROW, L_ROW2, L_HDR)]:
        if rows:
            section_hdr(ws4, ws4.max_row+1, f'-- {vendor} --', NC4, hfill)
            for i, r in enumerate(sorted(rows, key=lambda x: (x['cc'], str(x.get('ref','')))), 1):
                ws4.append([r['vendor'],r['cc'],r['ref'],r['inv5'],r['sap_monto'],r.get('texto_cab','')])
                brow(ws4, ws4.max_row, NC4, ra if i%2==0 else rb)
    tot_row(ws4, ws4.max_row+1, NC4, [5])
    col_w(ws4, [10, 16, 22, 12, 14, 35])

    # SIR SIN FACTURA
    ws5 = wb.create_sheet('SIR sin Factura'); NC5 = 6
    title(ws5, f'REGISTROS SIR SIN NUMERO DE FACTURA — {period.upper()}', NC5, '595959')
    ws5.append(['Proveedor','Centro de Costo','Fecha','Codigo Compras','Codigo Dev.','Monto ($)'])
    hdr(ws5, 2, NC5)
    for i, r in enumerate(sorted(p_nofac + l_nofac, key=lambda x: (x['cc'], x.get('fecha',''))), 3):
        monto = r.get('devolucion') if r.get('devolucion') else r.get('total')
        ws5.append([r['vendor'],r['cc'],r.get('fecha',''),r.get('cod',''),r.get('cod_dev',''),monto])
        brow(ws5, ws5.max_row, NC5, GRAY_F if i%2==0 else mk_fill('FFFFFF'))
    col_w(ws5, [10, 16, 12, 22, 22, 14])

    # TOTALES POR TIENDA — SAP vs PEPSI
    if store_totals:
        ws6 = wb.create_sheet('Totales x Tienda SAP-Pepsi'); NC6 = 4
        title(ws6, f'TOTALIZACION DE COMPRAS POR TIENDA — SAP vs PEPSI — {period.upper()}', NC6, '1F5C99')
        ws6.append(['Tienda', 'Total SAP ($)', 'Total Pepsi ($)', 'Diferencia ($)'])
        hdr(ws6, 2, NC6)
        for i, r in enumerate(store_totals, 3):
            ws6.append([r['cc'], r['total_sap_pepsi'], r['total_pepsi'], r['diferencia_pepsi']])
            fill = ERR_F if r['con_diferencia_pepsi'] else OK_F
            brow(ws6, ws6.max_row, NC6, fill)
            ws6.cell(ws6.max_row, 4).font = mk_font(size=10, bold=r['con_diferencia_pepsi'])
        tot_row(ws6, ws6.max_row + 1, NC6, [2, 3, 4])
        col_w(ws6, [14, 16, 16, 16])

        # TOTALES POR TIENDA — SAP vs LARKIN
        ws7 = wb.create_sheet('Totales x Tienda SAP-Larkin'); NC7 = 4
        title(ws7, f'TOTALIZACION DE COMPRAS POR TIENDA — SAP vs LARKIN — {period.upper()}', NC7, '7B2D8B')
        ws7.append(['Tienda', 'Total SAP ($)', 'Total Larkin ($)', 'Diferencia ($)'])
        hdr(ws7, 2, NC7)
        for i, r in enumerate(store_totals, 3):
            ws7.append([r['cc'], r['total_sap_larkin'], r['total_larkin'], r['diferencia_larkin']])
            fill = ERR_F if r['con_diferencia_larkin'] else OK_F
            brow(ws7, ws7.max_row, NC7, fill)
            ws7.cell(ws7.max_row, 4).font = mk_font(size=10, bold=r['con_diferencia_larkin'])
        tot_row(ws7, ws7.max_row + 1, NC7, [2, 3, 4])
        col_w(ws7, [14, 16, 16, 16])

    wb.save(output_path)
